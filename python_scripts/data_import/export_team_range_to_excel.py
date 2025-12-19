#!/usr/bin/env python3
"""
LoneStar Database to Excel - BY TEAM ID RANGE
==============================================

Exports only schedules for teams in a specific ID range.
Use after each scraping batch to create separate Excel files.

Usage:
    python export_team_range_to_excel.py 1 2400
    python export_team_range_to_excel.py 2401 3400
"""

import sys
import pyodbc
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

DB_CONNECTION_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=McKnights-PC\\SQLEXPRESS01;"
    "DATABASE=hs_football_database;"
    "Trusted_Connection=yes;"
)

EXCEL_BASE = r"C:\Users\demck\OneDrive\Football_2024\static-football-rankings\excel_files\HSF Texas 2025"
SHEET_NAME = "Lonestar"

# ============================================================================
# GAME PARSING
# ============================================================================

def parse_schedule_to_games(raw_text: str, team_name: str, season: int) -> list:
    """
    Parse raw schedule into LoneStar format - exactly as it appears on website
    
    Returns list of tuples: (week, team1_raw, score1, team2_raw, score2)
    
    Example line: "1  3Hedleyv 52 LFort Elliottv 34"
    Returns: ("1", "3Hedleyv", 52, "LFort Elliottv", 34)
    """
    
    games = []
    
    if not raw_text or not raw_text.strip():
        return games
    
    lines = raw_text.split('\n')
    
    for line in lines:
        line = line.strip()
        
        if not line or 'Record:' in line or 'WK' in line or 'Team SC' in line:
            continue
        
        if not any(char.isdigit() for char in line):
            continue
        
        try:
            # Save original line
            original_line = line
            
            # Extract week number at the very beginning
            # Handles: "1 ", "9/19 ", "10/31 ", "0 "
            week_match = re.match(r'^(\d+/?[\d]*)\s+', original_line)
            if week_match:
                week = week_match.group(1)
                # Remove week from line
                line = original_line[len(week_match.group(0)):].strip()
            else:
                week = ''
            
            # Now we need to carefully parse: Team1 Score1 Team2 Score2
            # The challenge: team names can have number prefixes (3Hedley)
            # Split by numbers with space before, space after is optional (for end of line)
            
            parts = re.split(r'\s+(\d+)(?:\s+|$)', line)
            
            # parts should be like:
            # ['3Hedleyv', '52', 'LFort Elliottv', '34', '']
            
            if len(parts) < 4:
                continue
            
            team1_raw = parts[0].strip()
            score1 = int(parts[1])
            team2_raw = parts[2].strip()
            score2 = int(parts[3])
            
            # Verify we got valid data
            if not team1_raw or not team2_raw:
                continue
            
            # Keep ALL junk characters intact
            games.append((week, team1_raw, score1, team2_raw, score2))
            
        except Exception as e:
            continue
    
    return games

# ============================================================================
# MAIN EXPORT
# ============================================================================

def main():
    # Get team ID range from command line
    if len(sys.argv) != 3:
        print("Usage: python export_team_range_to_excel.py START_ID END_ID")
        print("Example: python export_team_range_to_excel.py 1 2400")
        sys.exit(1)
    
    try:
        start_id = int(sys.argv[1])
        end_id = int(sys.argv[2])
    except ValueError:
        print("ERROR: Team IDs must be numbers")
        sys.exit(1)
    
    print("="*80)
    print(f"LoneStar Export: Team IDs {start_id} - {end_id}")
    print("="*80)
    print()
    
    # Generate output filename
    excel_file = f"{EXCEL_BASE}_teams_{start_id}-{end_id}.xlsx"
    
    # Connect to database
    print("Connecting to database...")
    conn = pyodbc.connect(DB_CONNECTION_STRING)
    cursor = conn.cursor()
    
    # Get schedules for team ID range
    print(f"Fetching schedules for team IDs {start_id} - {end_id}...")
    sql = """
        SELECT team_id, team_name, season, season_url, raw_schedule_text
        FROM lonestar_raw_schedules
        WHERE team_id >= ? AND team_id <= ?
        ORDER BY team_id, season DESC
    """
    
    cursor.execute(sql, start_id, end_id)
    rows = cursor.fetchall()
    
    print(f"✓ Found {len(rows)} schedules")
    print()
    
    if not rows:
        print("No schedules found in this team ID range!")
        return
    
    # Create new workbook
    print(f"Creating Excel file: {excel_file}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    
    # Write headers - exact LoneStar format
    headers = ['team_id', 'team_name', 'season', 'WK', 'Team', 'SC', 'Team', 'SC']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Process schedules
    print("Processing schedules...")
    
    current_row = 2
    total_games = 0
    schedules_processed = 0
    teams_processed = set()
    
    for schedule in rows:
        team_id, team_name, season, season_url, raw_text = schedule
        teams_processed.add(team_id)
        
        games = parse_schedule_to_games(raw_text, team_name, season)
        
        if not games:
            # Skip schedules with no games
            continue
        else:
            for week, team1_raw, score1, team2_raw, score2 in games:
                ws.cell(row=current_row, column=1, value=team_id)
                ws.cell(row=current_row, column=2, value=team_name)
                ws.cell(row=current_row, column=3, value=season)
                ws.cell(row=current_row, column=4, value=week)
                ws.cell(row=current_row, column=5, value=team1_raw)
                ws.cell(row=current_row, column=6, value=score1)
                ws.cell(row=current_row, column=7, value=team2_raw)
                ws.cell(row=current_row, column=8, value=score2)
                current_row += 1
                total_games += 1
        
        schedules_processed += 1
        
        if schedules_processed % 500 == 0:
            print(f"  Processed {schedules_processed}/{len(rows)} schedules...")
    
    print(f"✓ Processed {schedules_processed} schedules from {len(teams_processed)} teams")
    print(f"✓ Extracted {total_games} individual games")
    print()
    
    # Save workbook
    print("Saving Excel file...")
    wb.save(excel_file)
    print(f"✓ Saved to: {excel_file}")
    
    print()
    print("="*80)
    print("EXPORT COMPLETE")
    print("="*80)
    print(f"File: {excel_file}")
    print(f"Team IDs: {start_id} - {end_id}")
    print(f"Teams: {len(teams_processed)}")
    print(f"Schedules: {schedules_processed}")
    print(f"Games: {total_games}")
    print(f"Rows: {current_row - 1}")
    print()
    print("Next steps:")
    print("1. Open Excel file")
    print("2. Apply your formulas to columns I+ to standardize team names")
    print("3. Run import to SQL")
    print("="*80)
    
    conn.close()

if __name__ == "__main__":
    main()

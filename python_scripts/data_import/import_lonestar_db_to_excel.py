#!/usr/bin/env python3
"""
LoneStar Database to Excel Importer
====================================

Reads raw schedules from lonestar_raw_schedules table,
splits them into individual game rows,
and writes directly to Excel workbook.

Target: Excel file specified in EXCEL_FILE constant below
Sheet: Lonestar
Columns: A-H (team_id, team_name, season, season_url, raw_schedule_text, score1, team2_raw, score2)
"""

import pyodbc
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

DB_CONNECTION_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=McKnights-PC\\SQLEXPRESS01;"
    "DATABASE=hs_football_database;"
    "Trusted_Connection=yes;"
)

EXCEL_FILE = r"C:\Users\demck\OneDrive\Football_2024\static-football-rankings\excel_files\HSF Texas 2025.xlsx"
SHEET_NAME = "Lonestar"

# Column mapping (0-based)
COL_TEAM_ID = 0          # A
COL_TEAM_NAME = 1        # B
COL_SEASON = 2           # C
COL_SEASON_URL = 3       # D
COL_RAW_SCHEDULE = 4     # E
COL_SCORE1 = 5           # F
COL_TEAM2_RAW = 6        # G
COL_SCORE2 = 7           # H

# ============================================================================
# GAME PARSING
# ============================================================================

def parse_schedule_to_games(raw_text: str, team_name: str, season: int) -> list:
    """
    Parse raw schedule text into individual game rows
    Each game becomes: [opponent_score, opponent_raw, our_score]
    
    Example input line:
    "1  3Hedleyv 52 LFort Elliottv 34"
    
    For Fort Elliott Cougars, this returns:
    - opponent_score: 52
    - opponent_raw: "3Hedleyv" (with ALL prefix/suffix junk - Excel formulas clean it)
    - our_score: 34
    
    We keep ALL junk characters (prefixes like "3", "L", "@" and suffixes like "v", "y", "f")
    because the Excel VLOOKUP formulas are designed to strip them.
    """
    
    games = []
    
    if not raw_text or not raw_text.strip():
        return games
    
    lines = raw_text.split('\n')
    
    # Extract the base team name (without location prefix)
    # "Fort Elliott Cougars" -> "Fort Elliott"
    base_team_name = team_name.split('(')[0].strip()
    # Remove common mascot names if present
    mascot_words = ['Cougars', 'Eagles', 'Panthers', 'Tigers', 'Bears', 'Bulldogs', 
                    'Lions', 'Wildcats', 'Warriors', 'Hawks', 'Mustangs', 'Knights',
                    'Bearcats', 'Indians', 'Trojans', 'Pirates', 'Vikings', 'Rams']
    for mascot in mascot_words:
        if base_team_name.endswith(' ' + mascot):
            base_team_name = base_team_name[:-len(mascot)-1].strip()
            break
    
    for line in lines:
        line = line.strip()
        
        # Skip header lines, empty lines
        if not line or 'Record:' in line or 'WK' in line or 'Team SC' in line:
            continue
        
        # Skip lines that don't have scores (no numbers)
        if not any(char.isdigit() for char in line):
            continue
        
        try:
            # Remove week number at start if present (e.g., "1 " or "9/19 ")
            cleaned_line = re.sub(r'^\d+/?\d*\s+', '', line)
            
            # Find all sequences of: [team name] [score]
            # Pattern: word characters (with prefixes) followed by space and number
            # Split by numbers to find team segments
            parts = re.split(r'\s+(\d+)\s+', cleaned_line)
            
            # parts will be like: ['3Hedleyv', '52', 'LFort Elliottv', '34', '']
            if len(parts) < 4:
                continue
            
            team1_raw = parts[0].strip()
            score1 = int(parts[1])
            team2_raw = parts[2].strip()
            score2 = int(parts[3])
            
            # For matching purposes, remove trailing single-character flags
            # But keep the original raw strings for output
            team1_for_matching = re.sub(r'[a-zA-Z\$\#\*\@]$', '', team1_raw).strip()
            team2_for_matching = re.sub(r'[a-zA-Z\$\#\*\@]$', '', team2_raw).strip()
            
            # Determine which team is the opponent (not our team)
            # Check if base_team_name appears in team1 or team2
            opponent_raw = None
            opponent_score = None
            our_score = None
            
            if base_team_name.lower() in team1_for_matching.lower():
                # Team 1 is us, Team 2 is opponent
                opponent_raw = team2_raw  # Keep ALL junk chars
                opponent_score = score2
                our_score = score1
            elif base_team_name.lower() in team2_for_matching.lower():
                # Team 2 is us, Team 1 is opponent  
                opponent_raw = team1_raw  # Keep ALL junk chars
                opponent_score = score1
                our_score = score2
            else:
                # Can't determine - use team2 as opponent by default
                opponent_raw = team2_raw  # Keep ALL junk chars
                opponent_score = score2
                our_score = score1
            
            # If we found a valid opponent name
            if opponent_raw and len(opponent_raw) > 2:
                games.append((opponent_score, opponent_raw, our_score))
            
        except Exception as e:
            # Skip malformed lines
            continue
    
    return games

# ============================================================================
# MAIN IMPORT
# ============================================================================

def main():
    print("="*80)
    print("LoneStar Database → Excel Importer")
    print("="*80)
    print()
    
    # Connect to database
    print("Connecting to database...")
    conn = pyodbc.connect(DB_CONNECTION_STRING)
    cursor = conn.cursor()
    
    # Get all raw schedules
    print("Fetching raw schedules...")
    sql = """
        SELECT team_id, team_name, season, season_url, raw_schedule_text
        FROM lonestar_raw_schedules
        ORDER BY team_id, season DESC
    """
    
    cursor.execute(sql)
    rows = cursor.fetchall()
    
    print(f"✓ Found {len(rows)} schedules")
    print()
    
    if not rows:
        print("No data to import!")
        return
    
    # Load Excel workbook
    print(f"Opening Excel file: {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"ERROR: Excel file not found!")
        print(f"Please create: {EXCEL_FILE}")
        return
    
    # Get or create sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        print(f"✓ Found sheet '{SHEET_NAME}'")
        # Clear existing data (keep headers)
        if ws.max_row > 1:
            print(f"  Clearing {ws.max_row - 1} existing rows...")
            ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        print(f"✓ Created sheet '{SHEET_NAME}'")
        
        # Write headers
        headers = ['team_id', 'team_name', 'season', 'season_url', 'raw_schedule_text', 
                   'score1', 'team2_raw', 'score2']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # Process each schedule and write games
    print()
    print("Processing schedules and writing to Excel...")
    
    current_row = 2  # Start after header
    total_games = 0
    schedules_processed = 0
    
    for schedule in rows:
        team_id, team_name, season, season_url, raw_text = schedule
        
        # Parse schedule into individual games
        games = parse_schedule_to_games(raw_text, team_name, season)
        
        if not games:
            # No games found, but still write the raw schedule row
            ws.cell(row=current_row, column=COL_TEAM_ID + 1, value=team_id)
            ws.cell(row=current_row, column=COL_TEAM_NAME + 1, value=team_name)
            ws.cell(row=current_row, column=COL_SEASON + 1, value=season)
            ws.cell(row=current_row, column=COL_SEASON_URL + 1, value=season_url)
            ws.cell(row=current_row, column=COL_RAW_SCHEDULE + 1, value=raw_text)
            current_row += 1
            continue
        
        # Write each game as a separate row
        for score1, team2_raw, score2 in games:
            ws.cell(row=current_row, column=COL_TEAM_ID + 1, value=team_id)
            ws.cell(row=current_row, column=COL_TEAM_NAME + 1, value=team_name)
            ws.cell(row=current_row, column=COL_SEASON + 1, value=season)
            ws.cell(row=current_row, column=COL_SEASON_URL + 1, value=season_url)
            ws.cell(row=current_row, column=COL_RAW_SCHEDULE + 1, value=raw_text)
            ws.cell(row=current_row, column=COL_SCORE1 + 1, value=score1)
            ws.cell(row=current_row, column=COL_TEAM2_RAW + 1, value=team2_raw)
            ws.cell(row=current_row, column=COL_SCORE2 + 1, value=score2)
            current_row += 1
            total_games += 1
        
        schedules_processed += 1
        
        # Progress indicator
        if schedules_processed % 100 == 0:
            print(f"  Processed {schedules_processed}/{len(rows)} schedules, {total_games} games...")
    
    print(f"✓ Processed {schedules_processed} schedules")
    print(f"✓ Extracted {total_games} individual games")
    print()
    
    # Save workbook
    print("Saving Excel file...")
    wb.save(EXCEL_FILE)
    print(f"✓ Saved to: {EXCEL_FILE}")
    
    print()
    print("="*80)
    print("IMPORT COMPLETE")
    print("="*80)
    print(f"Sheet: {SHEET_NAME}")
    print(f"Rows: {current_row - 1} (excluding header)")
    print(f"Games: {total_games}")
    print()
    print("Next steps:")
    print("1. Open Excel file and verify data")
    print("2. Apply your formulas to parse team names and standardize")
    print("3. Run the Python import to SQL")
    print("="*80)
    
    conn.close()

if __name__ == "__main__":
    main()
